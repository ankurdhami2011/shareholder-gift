from django.contrib import messages
from django.contrib.auth.hashers import check_password
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from accounts.permission_utils import permission_required
from django.db import transaction
from django.db.models import Count, Q
from django.core.paginator import Paginator
from django.http import FileResponse, Http404, HttpResponse
from masters.models import Branch
from gift_requests.models import GiftCycle
from django.db.models import Exists, OuterRef
from django.conf import settings
import uuid
from .forms import StaffLoginForm
from .models import StaffUser
import os
import mimetypes
from shares.models import Share
from shareholders.models import Shareholder
from django.views.decorators.clickjacking import xframe_options_sameorigin
from gift_requests.models import GiftRequestDocument
from gift_requests.models import GiftRequest, GiftRequestStatusHistory
from audit_logs.utils import create_audit_log
from notifications.utils import (
    send_request_accepted_sms,
    send_request_rejected_sms,
    send_tracking_created_sms,
    send_request_delivered_sms,
)



def get_logged_in_staff(request):
    staff_id = request.session.get('staff_user_id')
    if not staff_id:
        return None
    return StaffUser.objects.filter(id=staff_id, is_active=True).first()


def staff_login_required(view_func):
    def wrapper(request, *args, **kwargs):
        staff = get_logged_in_staff(request)
        if not staff:
            return redirect('staff-web-login')
        request.staff_user = staff
        return view_func(request, *args, **kwargs)
    return wrapper

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font


MAX_EXCEL_FILE_SIZE = 5 * 1024 * 1024   # 5 MB
MAX_DOCUMENT_FILE_SIZE = 10 * 1024 * 1024   # 10 MB

ALLOWED_EXCEL_EXTENSIONS = ['.xlsx']
ALLOWED_EXCEL_CONTENT_TYPES = [
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/octet-stream',
]

ALLOWED_DOCUMENT_EXTENSIONS = ['.pdf', '.jpg', '.jpeg', '.png']
ALLOWED_DOCUMENT_CONTENT_TYPES = [
    'application/pdf',
    'image/jpeg',
    'image/png',
    'application/octet-stream',
]


def validate_excel_upload(uploaded_file):
    if not uploaded_file:
        return 'Please select an Excel file'

    file_name = uploaded_file.name.lower()
    file_ext = os.path.splitext(file_name)[1]

    if file_ext not in ALLOWED_EXCEL_EXTENSIONS:
        return 'Only .xlsx files are allowed'

    if uploaded_file.size > MAX_EXCEL_FILE_SIZE:
        return 'Excel file size must be 5 MB or less'

    content_type = getattr(uploaded_file, 'content_type', '') or ''
    if content_type and content_type not in ALLOWED_EXCEL_CONTENT_TYPES:
        return f'Invalid Excel content type: {content_type}'

    return None


def validate_document_upload(uploaded_file):
    if not uploaded_file:
        return 'Invalid document'

    file_name = uploaded_file.name.lower()
    file_ext = os.path.splitext(file_name)[1]

    if file_ext not in ALLOWED_DOCUMENT_EXTENSIONS:
        return 'Only PDF, JPG, JPEG, PNG files are allowed'

    if uploaded_file.size > MAX_DOCUMENT_FILE_SIZE:
        return 'Document file size must be 10 MB or less'

    content_type = getattr(uploaded_file, 'content_type', '') or ''
    if content_type and content_type not in ALLOWED_DOCUMENT_CONTENT_TYPES:
        return f'Invalid document content type: {content_type}'

    return None


def validate_excel_headers_and_rows(ws, expected_headers):
    headers_found = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in ws[1]
    ]

    if headers_found != expected_headers:
        return headers_found, f'Invalid headers. Expected: {expected_headers}. Found: {headers_found}'

    if ws.max_row < 2:
        return headers_found, 'Excel file has no data rows'

    return headers_found, None
    
@staff_login_required
@permission_required('can_view_reports')
def staff_report_summary(request):
    staff = request.staff_user

    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()
    branch_id = request.GET.get('branch_id', '').strip()
    status_filter = request.GET.get('status', '').strip()
    request_no = request.GET.get('request_no', '').strip()
    per_page = request.GET.get('per_page', '10').strip()

    queryset = GiftRequest.objects.select_related(
        'branch', 'shareholder', 'share', 'gift_cycle'
    ).all().order_by('-id')

    # Branch access filter
    if not (staff.has_all_branch_access or staff.role_type in ['MASTER_STAFF', 'ADMIN']):
        allowed_branch_ids = list(staff.branch_accesses.values_list('branch_id', flat=True))
        queryset = queryset.filter(branch_id__in=allowed_branch_ids)
        branch_queryset = Branch.objects.filter(id__in=allowed_branch_ids, is_active=True)
    else:
        branch_queryset = Branch.objects.filter(is_active=True)

    # Date filters
    if from_date:
        queryset = queryset.filter(submitted_at__date__gte=from_date)

    if to_date:
        queryset = queryset.filter(submitted_at__date__lte=to_date)

    # Branch filter
    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    # Status filter
    if status_filter:
        queryset = queryset.filter(request_status=status_filter)

    # Request no filter
    if request_no:
        queryset = queryset.filter(request_no__icontains=request_no)

    # Top cards summary
    status_summary = {
        'inserted': queryset.count(),
        'accepted': queryset.filter(request_status='ACCEPTED').count(),
        'rejected': queryset.filter(request_status='REJECTED').count(),
        'shipped': queryset.filter(request_status='SHIPPED').count(),
        'delivered': queryset.filter(request_status='DELIVERED').count(),
    }

    # Branch-wise summary
    branch_summary = (
        queryset.values('branch__branch_name')
        .annotate(
            inserted=Count('id'),
            accepted=Count('id', filter=Q(request_status='ACCEPTED')),
            rejected=Count('id', filter=Q(request_status='REJECTED')),
            shipped=Count('id', filter=Q(request_status='SHIPPED')),
            delivered=Count('id', filter=Q(request_status='DELIVERED')),
        )
        .order_by('branch__branch_name')
    )

    # Pagination
    if per_page not in ['10', '20']:
        per_page = '10'

    paginator = Paginator(queryset, int(per_page))
    page_number = request.GET.get('page')
    details = paginator.get_page(page_number)

    context = {
        'staff_user': staff,
        'status_summary': status_summary,
        'branch_summary': branch_summary,
        'branch_queryset': branch_queryset,
        'from_date': from_date,
        'to_date': to_date,
        'branch_id': branch_id,
        'status_filter': status_filter,
        'request_no': request_no,
        'per_page': per_page,
        'details': details,
        'status_choices': [
            ('', 'All Status'),
            ('PENDING', 'Pending'),
            ('ACCEPTED', 'Accepted'),
            ('REJECTED', 'Rejected'),
            ('SHIPPED', 'Shipped'),
            ('DELIVERED', 'Delivered'),
        ],
    }

    return render(request, 'staff/report_summary.html', context)
@staff_login_required
@permission_required('can_view_reports')
def staff_report_export_excel(request):
    staff = request.staff_user

    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()
    branch_id = request.GET.get('branch_id', '').strip()
    status_filter = request.GET.get('status', '').strip()
    request_no = request.GET.get('request_no', '').strip()

    queryset = GiftRequest.objects.select_related(
        'branch', 'shareholder', 'share', 'gift_cycle'
    ).all().order_by('-id')

    if not (staff.has_all_branch_access or staff.role_type in ['MASTER_STAFF', 'ADMIN']):
        allowed_branch_ids = list(staff.branch_accesses.values_list('branch_id', flat=True))
        queryset = queryset.filter(branch_id__in=allowed_branch_ids)

    if from_date:
        queryset = queryset.filter(submitted_at__date__gte=from_date)

    if to_date:
        queryset = queryset.filter(submitted_at__date__lte=to_date)

    if branch_id:
        queryset = queryset.filter(branch_id=branch_id)

    if status_filter:
        queryset = queryset.filter(request_status=status_filter)

    if request_no:
        queryset = queryset.filter(request_no__icontains=request_no)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ['ID', 'Request No', 'Date', 'Branch', 'Shareholder', 'Folio No', 'Gift Item', 'Status']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in queryset:
        shareholder_name = getattr(row.shareholder, 'shareholder_name', str(row.shareholder))
        folio_no = getattr(row.shareholder, 'folio_no', '-')
        gift_item = getattr(row.gift_cycle, 'gift_name', str(row.gift_cycle))

        ws.append([
            row.id,
            row.request_no,
            row.submitted_at.strftime('%d-%m-%Y') if row.submitted_at else '',
            row.branch.branch_name if row.branch else '',
            shareholder_name,
            folio_no,
            gift_item,
            row.request_status,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=staff_report.xlsx'
    wb.save(response)
    return response    
def staff_has_branch_access(staff_user, branch_id):
    if staff_user.has_all_branch_access or staff_user.role_type in ['MASTER_STAFF', 'ADMIN']:
        return True
    return staff_user.branch_accesses.filter(branch_id=branch_id).exists()


from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User

def staff_web_login(request):
    if request.session.get('staff_user_id'):
        return redirect('staff-web-dashboard')

    form = StaffLoginForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        username = form.cleaned_data['username']
        password = form.cleaned_data['password']

        # authenticate using Django
        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, 'Invalid username or password')
        else:
            # map to staff
            staff = StaffUser.objects.filter(username=username).first()
            
            if not staff:
                staff = StaffUser.objects.create(
                    username=username,
                    staff_code=f"STF{user.id:03}",
                    full_name=f"{user.first_name} {user.last_name}".strip() or username,
                    mobile_number='',
                    email=user.email,
                    password_hash=user.password,
                    role_type='ADMIN',
                    has_all_branch_access=True,
                    is_active=True,
                )
            elif not staff.is_active:
                messages.error(request, 'Staff account is inactive')
                return render(request, 'staff/login.html', {'form': form})
            else:
                if staff.password_hash != user.password:
                    staff.password_hash = user.password
                    staff.save(update_fields=['password_hash'])

            login(request, user)

            request.session['staff_user_id'] = staff.id
            staff.last_login_at = timezone.now()
            staff.save(update_fields=['last_login_at'])

            return redirect('staff-web-dashboard')

    return render(request, 'staff/login.html', {'form': form})


from django.contrib.auth import logout

def staff_web_logout(request):
    logout(request)  # 🔥 add this
    request.session.flush()
    return redirect('staff-web-login')
    
    
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.utils import timezone
import json


@staff_login_required
def staff_dashboard(request):
    staff = request.staff_user

    request_queryset = GiftRequest.objects.select_related('branch').all()
    share_queryset = Share.objects.select_related('branch').filter(is_active=True)

    if not (staff.has_all_branch_access or staff.role_type in ['MASTER_STAFF', 'ADMIN']):
        allowed_branch_ids = list(staff.branch_accesses.values_list('branch_id', flat=True))
        request_queryset = request_queryset.filter(branch_id__in=allowed_branch_ids)
        share_queryset = share_queryset.filter(branch_id__in=allowed_branch_ids)

    today = timezone.localdate()
    last_7_days = today - timezone.timedelta(days=6)

    daily_requests_qs = (
        request_queryset
        .filter(submitted_at__date__gte=last_7_days)
        .annotate(day=TruncDate('submitted_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    
    daily_requests_map = {str(x['day']): x['count'] for x in daily_requests_qs}
    
    request_dates = []
    request_counts = []
    for i in range(7):
        day = last_7_days + timezone.timedelta(days=i)
        day_str = str(day)
        request_dates.append(day_str)
        request_counts.append(daily_requests_map.get(day_str, 0))
    
    
    daily_delivered_qs = (
        request_queryset
        .filter(delivered_at__isnull=False, delivered_at__date__gte=last_7_days)
        .annotate(day=TruncDate('delivered_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    
    daily_delivered_map = {str(x['day']): x['count'] for x in daily_delivered_qs}
    
    delivered_dates = []
    delivered_counts = []
    for i in range(7):
        day = last_7_days + timezone.timedelta(days=i)
        day_str = str(day)
        delivered_dates.append(day_str)
        delivered_counts.append(daily_delivered_map.get(day_str, 0))

    # Branch-wise Requests
    branch_data = (
        request_queryset
        .values('branch__branch_name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    branch_labels = [x['branch__branch_name'] for x in branch_data]
    branch_counts = [x['count'] for x in branch_data]

    # Share Status Distribution
    share_status_data = (
        share_queryset
        .values('gift_status')
        .annotate(count=Count('id'))
    )

    share_status_labels = [x['gift_status'] for x in share_status_data]
    share_status_counts = [x['count'] for x in share_status_data]

    context = {
        'pending_count': request_queryset.filter(request_status='PENDING').count(),
        'accepted_count': request_queryset.filter(request_status='ACCEPTED').count(),
        'rejected_count': request_queryset.filter(request_status='REJECTED').count(),
        'shipped_count': request_queryset.filter(request_status='SHIPPED').count(),
        'delivered_count': request_queryset.filter(request_status='DELIVERED').count(),
        'total_request_count': request_queryset.count(),
    
        'total_share_count': share_queryset.count(),
        'eligible_share_count': share_queryset.filter(gift_status='ELIGIBLE').count(),
        'stopped_share_count': share_queryset.filter(gift_status='STOPPED').count(),
    
        'today_request_count': request_queryset.filter(submitted_at__date=today).count(),
        'today_delivered_count': request_queryset.filter(delivered_at__date=today).count(),
    
        'recent_requests': request_queryset.select_related(
            'shareholder', 'share', 'branch'
        ).order_by('-id')[:10],
    
        'request_dates': json.dumps(request_dates),
        'request_counts': json.dumps(request_counts),
        'delivered_dates': json.dumps(delivered_dates),
        'delivered_counts': json.dumps(delivered_counts),
        'branch_labels': json.dumps(branch_labels),
        'branch_counts': json.dumps(branch_counts),
        'share_status_labels': json.dumps(share_status_labels),
        'share_status_counts': json.dumps(share_status_counts),
    
        'notification_count': 0,
        'staff_user': staff,
    }
    return render(request, 'staff/dashboard.html', context)

@staff_login_required
@permission_required('can_view_requests')
def staff_request_list(request):
    staff = request.staff_user

    queryset = GiftRequest.objects.select_related(
        'shareholder', 'share', 'branch', 'gift_cycle'
    ).order_by('-id')

    if not (staff.has_all_branch_access or staff.role_type in ['MASTER_STAFF', 'ADMIN']):
        allowed_branch_ids = list(staff.branch_accesses.values_list('branch_id', flat=True))
        queryset = queryset.filter(branch_id__in=allowed_branch_ids)

    request_no = request.GET.get('request_no', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if request_no:
        queryset = queryset.filter(request_no__icontains=request_no)

    if status_filter:
        queryset = queryset.filter(request_status=status_filter)

    return render(request, 'staff/request_list.html', {
        'items': queryset,
        'request_no': request_no,
        'status_filter': status_filter,
        'staff_user': staff,
    })


@staff_login_required
@permission_required('can_view_requests')
def staff_request_detail(request, gift_request_id):
    staff = request.staff_user

    obj = get_object_or_404(
        GiftRequest.objects.select_related(
            'shareholder', 'share', 'branch', 'gift_cycle'
        ).prefetch_related('documents', 'status_history'),
        id=gift_request_id
    )

    if not staff_has_branch_access(staff, obj.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-request-list')

    return render(request, 'staff/request_detail.html', {
        'item': obj,
        'staff_user': staff,
    })


@staff_login_required
@permission_required('can_accept_request')
def staff_accept_request(request, gift_request_id):
    staff = request.staff_user
    obj = get_object_or_404(GiftRequest, id=gift_request_id)

    if not staff_has_branch_access(staff, obj.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-request-list')

    if obj.request_status != 'PENDING':
        messages.error(request, 'Only pending requests can be accepted')
        return redirect('staff-web-request-detail', gift_request_id=obj.id)

    remarks = request.POST.get('remarks', '').strip()

    with transaction.atomic():
        old_status = obj.request_status
        old_accepted_at = obj.accepted_at
        old_updated_by_type = obj.updated_by_type
        old_updated_by_id = obj.updated_by_id

        obj.request_status = 'ACCEPTED'
        obj.accepted_at = timezone.now()
        obj.updated_by_type = 'STAFF'
        obj.updated_by_id = staff.id
        obj.save()

        GiftRequestStatusHistory.objects.create(
            gift_request=obj,
            old_status=old_status,
            new_status='ACCEPTED',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            remarks=remarks or None
        )

        create_audit_log(
            module_name='Gift Request',
            table_name='gift_requests_giftrequest',
            record_id=obj.id,
            action_type='ACCEPT_REQUEST',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            branch=obj.branch,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            remarks=remarks or 'Request accepted by staff',
            field_changes=[
                {'column_name': 'request_status', 'old_value': old_status, 'new_value': obj.request_status},
                {'column_name': 'accepted_at', 'old_value': old_accepted_at, 'new_value': obj.accepted_at},
                {'column_name': 'updated_by_type', 'old_value': old_updated_by_type, 'new_value': obj.updated_by_type},
                {'column_name': 'updated_by_id', 'old_value': old_updated_by_id, 'new_value': obj.updated_by_id},
            ]
        )

        send_request_accepted_sms(obj)

    messages.success(request, 'Request accepted successfully')
    return redirect('staff-web-request-detail', gift_request_id=obj.id)


@staff_login_required
@permission_required('can_reject_request')
def staff_reject_request(request, gift_request_id):
    staff = request.staff_user
    obj = get_object_or_404(GiftRequest, id=gift_request_id)

    if not staff_has_branch_access(staff, obj.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-request-list')

    if obj.request_status != 'PENDING':
        messages.error(request, 'Only pending requests can be rejected')
        return redirect('staff-web-request-detail', gift_request_id=obj.id)

    rejection_reason = request.POST.get('rejection_reason', '').strip()
    remarks = request.POST.get('remarks', '').strip()

    if not rejection_reason:
        messages.error(request, 'Rejection reason is required')
        return redirect('staff-web-request-detail', gift_request_id=obj.id)

    with transaction.atomic():
        old_status = obj.request_status
        old_rejection_reason = obj.rejection_reason
        old_rejected_at = obj.rejected_at
        old_updated_by_type = obj.updated_by_type
        old_updated_by_id = obj.updated_by_id

        obj.request_status = 'REJECTED'
        obj.rejection_reason = rejection_reason
        obj.rejected_at = timezone.now()
        obj.updated_by_type = 'STAFF'
        obj.updated_by_id = staff.id
        obj.save()

        GiftRequestStatusHistory.objects.create(
            gift_request=obj,
            old_status=old_status,
            new_status='REJECTED',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            remarks=remarks or None
        )

        create_audit_log(
            module_name='Gift Request',
            table_name='gift_requests_giftrequest',
            record_id=obj.id,
            action_type='REJECT_REQUEST',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            branch=obj.branch,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            remarks=remarks or 'Request rejected by staff',
            field_changes=[
                {'column_name': 'request_status', 'old_value': old_status, 'new_value': obj.request_status},
                {'column_name': 'rejection_reason', 'old_value': old_rejection_reason, 'new_value': obj.rejection_reason},
                {'column_name': 'rejected_at', 'old_value': old_rejected_at, 'new_value': obj.rejected_at},
                {'column_name': 'updated_by_type', 'old_value': old_updated_by_type, 'new_value': obj.updated_by_type},
                {'column_name': 'updated_by_id', 'old_value': old_updated_by_id, 'new_value': obj.updated_by_id},
            ]
        )

        send_request_rejected_sms(obj)

    messages.success(request, 'Request rejected successfully')
    return redirect('staff-web-request-detail', gift_request_id=obj.id)

@staff_login_required
@permission_required('can_ship_request')
def staff_ship_request(request, gift_request_id):
    staff = request.staff_user
    obj = get_object_or_404(GiftRequest, id=gift_request_id)

    if not staff_has_branch_access(staff, obj.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-request-list')

    if obj.request_status != 'ACCEPTED':
        messages.error(request, 'Only accepted requests can be shipped')
        return redirect('staff-web-request-detail', gift_request_id=obj.id)

    courier_name = request.POST.get('courier_name', '').strip()
    tracking_number = request.POST.get('tracking_number', '').strip()
    remarks = request.POST.get('remarks', '').strip()

    if not courier_name or not tracking_number:
        messages.error(request, 'Courier name and tracking number are required')
        return redirect('staff-web-request-detail', gift_request_id=obj.id)

    with transaction.atomic():
        old_status = obj.request_status
        old_courier_name = obj.courier_name
        old_tracking_number = obj.tracking_number
        old_shipped_at = obj.shipped_at
        old_updated_by_type = obj.updated_by_type
        old_updated_by_id = obj.updated_by_id

        obj.request_status = 'SHIPPED'
        obj.courier_name = courier_name
        obj.tracking_number = tracking_number
        obj.shipped_at = timezone.now()
        obj.updated_by_type = 'STAFF'
        obj.updated_by_id = staff.id
        obj.save()

        GiftRequestStatusHistory.objects.create(
            gift_request=obj,
            old_status=old_status,
            new_status='SHIPPED',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            remarks=remarks or None
        )

        create_audit_log(
            module_name='Gift Request',
            table_name='gift_requests_giftrequest',
            record_id=obj.id,
            action_type='SHIP_REQUEST',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            branch=obj.branch,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            remarks=remarks or 'Request shipped by staff',
            field_changes=[
                {'column_name': 'request_status', 'old_value': old_status, 'new_value': obj.request_status},
                {'column_name': 'courier_name', 'old_value': old_courier_name, 'new_value': obj.courier_name},
                {'column_name': 'tracking_number', 'old_value': old_tracking_number, 'new_value': obj.tracking_number},
                {'column_name': 'shipped_at', 'old_value': old_shipped_at, 'new_value': obj.shipped_at},
                {'column_name': 'updated_by_type', 'old_value': old_updated_by_type, 'new_value': obj.updated_by_type},
                {'column_name': 'updated_by_id', 'old_value': old_updated_by_id, 'new_value': obj.updated_by_id},
            ]
        )

        send_tracking_created_sms(obj)

    messages.success(request, 'Tracking details updated successfully')
    return redirect('staff-web-request-detail', gift_request_id=obj.id)


@staff_login_required
@permission_required('can_deliver_request')
def staff_deliver_request(request, gift_request_id):
    staff = request.staff_user
    obj = get_object_or_404(GiftRequest, id=gift_request_id)

    if not staff_has_branch_access(staff, obj.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-request-list')

    if obj.request_status != 'SHIPPED':
        messages.error(request, 'Only shipped requests can be marked delivered')
        return redirect('staff-web-request-detail', gift_request_id=obj.id)

    remarks = request.POST.get('remarks', '').strip()

    with transaction.atomic():
        old_status = obj.request_status
        old_delivered_at = obj.delivered_at
        old_updated_by_type = obj.updated_by_type
        old_updated_by_id = obj.updated_by_id

        obj.request_status = 'DELIVERED'
        obj.delivered_at = timezone.now()
        obj.updated_by_type = 'STAFF'
        obj.updated_by_id = staff.id
        obj.save()

        GiftRequestStatusHistory.objects.create(
            gift_request=obj,
            old_status=old_status,
            new_status='DELIVERED',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            remarks=remarks or None
        )

        create_audit_log(
            module_name='Gift Request',
            table_name='gift_requests_giftrequest',
            record_id=obj.id,
            action_type='DELIVER_REQUEST',
            changed_by_type='STAFF',
            changed_by_id=staff.id,
            branch=obj.branch,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            remarks=remarks or 'Request marked delivered by staff',
            field_changes=[
                {'column_name': 'request_status', 'old_value': old_status, 'new_value': obj.request_status},
                {'column_name': 'delivered_at', 'old_value': old_delivered_at, 'new_value': obj.delivered_at},
                {'column_name': 'updated_by_type', 'old_value': old_updated_by_type, 'new_value': obj.updated_by_type},
                {'column_name': 'updated_by_id', 'old_value': old_updated_by_id, 'new_value': obj.updated_by_id},
            ]
        )
        send_request_delivered_sms(obj)
    messages.success(request, 'Request marked as delivered successfully')
    return redirect('staff-web-request-detail', gift_request_id=obj.id)
    
@staff_login_required
@permission_required('can_view_requests')
def staff_document_download(request, document_id):
    staff = request.staff_user

    doc = get_object_or_404(
        GiftRequestDocument.objects.select_related('gift_request__branch'),
        id=document_id
    )

    if not staff_has_branch_access(staff, doc.gift_request.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-request-list')

    if not doc.file_path or not os.path.exists(doc.file_path):
        raise Http404("Document file not found")

    content_type, _ = mimetypes.guess_type(doc.file_path)
    if not content_type:
        content_type = doc.mime_type or 'application/octet-stream'

    response = FileResponse(open(doc.file_path, 'rb'), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{doc.original_file_name}"'
    return response

@staff_login_required
@permission_required('can_view_requests')
@xframe_options_sameorigin
def staff_document_preview(request, document_id):
    staff = request.staff_user

    doc = get_object_or_404(
        GiftRequestDocument.objects.select_related('gift_request__branch'),
        id=document_id
    )

    if not staff_has_branch_access(staff, doc.gift_request.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-request-list')

    if not doc.file_path or not os.path.exists(doc.file_path):
        raise Http404("Document file not found")

    content_type, _ = mimetypes.guess_type(doc.file_path)
    if not content_type:
        content_type = doc.mime_type or 'application/octet-stream'

    response = FileResponse(open(doc.file_path, 'rb'), content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{doc.original_file_name}"'
    return response
    
@staff_login_required
@permission_required('can_manage_share_status')
def staff_share_status_list(request):
    staff = request.staff_user

    search = request.GET.get('search', '').strip()

    queryset = Share.objects.select_related(
        'shareholder', 'branch'
    ).order_by('-id')

    if not (staff.has_all_branch_access or staff.role_type in ['MASTER_STAFF', 'ADMIN']):
        allowed_branch_ids = list(staff.branch_accesses.values_list('branch_id', flat=True))
        queryset = queryset.filter(branch_id__in=allowed_branch_ids)

    if search:
        queryset = queryset.filter(
            Q(share_number__icontains=search) |
            Q(shareholder__shareholder_name__icontains=search) |
            Q(shareholder__shareholder_code__icontains=search)
        )

    return render(request, 'staff/share_status_list.html', {
        'items': queryset[:100],
        'search': search,
        'staff_user': staff,
    })


@staff_login_required
@permission_required('can_manage_share_status')
def staff_share_status_update(request, share_id):
    staff = request.staff_user

    share = get_object_or_404(
        Share.objects.select_related('branch', 'shareholder'),
        id=share_id
    )

    if not staff_has_branch_access(staff, share.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-share-status-list')

    if request.method == 'POST':
        new_status = request.POST.get('new_status', '').strip()
        remarks = request.POST.get('remarks', '').strip()

        if new_status not in ['ELIGIBLE', 'STOPPED']:
            messages.error(request, 'Invalid status selected')
            return redirect('staff-web-share-status-list')

        if not remarks:
            messages.error(request, 'Remarks is required')
            return redirect('staff-web-share-status-list')

        old_status = share.gift_status

        if old_status == new_status:
            messages.error(request, 'No change in status')
            return redirect('staff-web-share-status-list')

        share.gift_status = new_status
        share.stop_reason = remarks
        share.save()

        messages.success(request, f'Share {share.share_number} updated successfully')
        return redirect('staff-web-share-status-list')

    return redirect('staff-web-share-status-list')
    
@staff_login_required
@permission_required('can_create_request')
def staff_create_request_search(request):
    staff = request.staff_user
    search = request.GET.get('search', '').strip()

    active_gift_cycle = GiftCycle.objects.filter(is_active=True).order_by('-id').first()

    queryset = Share.objects.select_related(
        'shareholder', 'branch'
    ).filter(
        is_active=True
    ).order_by('-id')

    if not (staff.has_all_branch_access or staff.role_type in ['MASTER_STAFF', 'ADMIN']):
        allowed_branch_ids = list(staff.branch_accesses.values_list('branch_id', flat=True))
        queryset = queryset.filter(branch_id__in=allowed_branch_ids)

    if search:
        queryset = queryset.filter(
            Q(share_number__icontains=search) |
            Q(shareholder__shareholder_code__icontains=search) |
            Q(shareholder__shareholder_name__icontains=search) |
            Q(shareholder__mobile_number__icontains=search)
        )

    items = []
    for share in queryset[:100]:
        last_request = None
        if active_gift_cycle:
            last_request = GiftRequest.objects.filter(
                share=share,
                gift_cycle=active_gift_cycle
            ).order_by('-id').first()

        has_blocking_request = False
        last_request_status = ''
        last_rejection_reason = ''

        if last_request:
            last_request_status = last_request.request_status or ''
            last_rejection_reason = last_request.rejection_reason or ''

            if last_request.request_status in ['PENDING', 'ACCEPTED', 'SHIPPED', 'DELIVERED']:
                has_blocking_request = True

        items.append({
            'id': share.id,
            'share_number': share.share_number,
            'shareholder_code': share.shareholder.shareholder_code,
            'shareholder_name': share.shareholder.shareholder_name,
            'mobile_number': share.shareholder.mobile_number,
            'branch_name': share.branch.branch_name,
            'gift_status': share.gift_status,
            'has_blocking_request': has_blocking_request,
            'last_request_status': last_request_status,
            'last_rejection_reason': last_rejection_reason,
        })

    return render(request, 'staff/create_request_search.html', {
        'items': items,
        'search': search,
        'staff_user': staff,
        'active_gift_cycle': active_gift_cycle,
    })
    
@staff_login_required
@permission_required('can_create_request')
def staff_create_request_form(request, share_id):
    staff = request.staff_user

    share = get_object_or_404(
        Share.objects.select_related('shareholder', 'branch'),
        id=share_id,
        is_active=True
    )

    if not staff_has_branch_access(staff, share.branch_id):
        messages.error(request, 'No branch access')
        return redirect('staff-web-create-request-search')

    if share.gift_status != 'ELIGIBLE':
        messages.error(request, 'Only ELIGIBLE shares can create request')
        return redirect('staff-web-create-request-search')

    if request.method == 'POST':
        mobile_number = request.POST.get('mobile_number', '').strip()
        recipient_name = request.POST.get('recipient_name', '').strip()
        address_line1 = request.POST.get('address_line1', '').strip()
        address_line2 = request.POST.get('address_line2', '').strip()
        city = request.POST.get('city', '').strip()
        state = request.POST.get('state', '').strip()
        pincode = request.POST.get('pincode', '').strip()
        customer_type = request.POST.get('customer_type', '').strip() or 'KEYPAD_PHONE'
        remarks = request.POST.get('remarks', '').strip()
        uploaded_files = request.FILES.getlist('documents')

        if not mobile_number:
            messages.error(request, 'Mobile number is required')
            return redirect('staff-web-create-request-form', share_id=share.id)

        if not recipient_name:
            messages.error(request, 'Recipient name is required')
            return redirect('staff-web-create-request-form', share_id=share.id)

        if not address_line1 or not city or not state or not pincode:
            messages.error(request, 'Address Line 1, City, State and Pincode are required')
            return redirect('staff-web-create-request-form', share_id=share.id)

        if not uploaded_files:
            messages.error(request, 'At least one document is required')
            return redirect('staff-web-create-request-form', share_id=share.id)

        for uploaded_file in uploaded_files:
            doc_error = validate_document_upload(uploaded_file)
            if doc_error:
                messages.error(request, f'{uploaded_file.name}: {doc_error}')
                return redirect('staff-web-create-request-form', share_id=share.id)

        try:
            with transaction.atomic():
                active_gift_cycle = GiftCycle.objects.filter(is_active=True).order_by('-id').first()

                if not active_gift_cycle:
                    messages.error(request, 'No active gift cycle found')
                    return redirect('staff-web-create-request-form', share_id=share.id)

                duplicate_exists = GiftRequest.objects.filter(
                    share=share,
                    gift_cycle=active_gift_cycle,
                    request_status__in=['PENDING', 'ACCEPTED', 'SHIPPED', 'DELIVERED']
                ).exists()

                if duplicate_exists:
                    messages.error(request, 'Active or completed request already exists for this share in current gift cycle')
                    return redirect('staff-web-create-request-search')

                shareholder = share.shareholder
                shareholder.mobile_number = mobile_number
                shareholder.address_line1 = address_line1
                shareholder.address_line2 = address_line2 or None
                shareholder.city = city
                shareholder.state = state
                shareholder.pincode = pincode
                shareholder.save()

                gift_request_field_names = {f.name for f in GiftRequest._meta.fields}

                request_no = f"GR-{timezone.now().strftime('%Y%m%d%H%M%S%f')}"

                create_kwargs = {
                    'request_no': request_no,
                    'shareholder': share.shareholder,
                    'share': share,
                    'branch': share.branch,
                    'gift_cycle': active_gift_cycle,
                    'mobile_number': mobile_number,
                    'request_status': 'PENDING',
                    'updated_by_type': 'STAFF',
                    'updated_by_id': staff.id,
                }

                if 'rejection_reason' in gift_request_field_names:
                    create_kwargs['rejection_reason'] = None

                if 'tracking_number' in gift_request_field_names:
                    create_kwargs['tracking_number'] = None

                if 'courier_name' in gift_request_field_names:
                    create_kwargs['courier_name'] = None

                if 'request_source' in gift_request_field_names:
                    create_kwargs['request_source'] = 'STAFF'

                if 'customer_type' in gift_request_field_names:
                    create_kwargs['customer_type'] = customer_type

                if 'remarks' in gift_request_field_names:
                    create_kwargs['remarks'] = remarks or None

                obj = GiftRequest.objects.create(**create_kwargs)

                GiftRequestStatusHistory.objects.create(
                    gift_request=obj,
                    old_status=None,
                    new_status='PENDING',
                    changed_by_type='STAFF',
                    changed_by_id=staff.id,
                    remarks=remarks or 'Request created by staff'
                )

                doc_field_names = {f.name for f in GiftRequestDocument._meta.fields}
                save_dir = os.path.join(settings.MEDIA_ROOT, 'gift_request_docs')
                os.makedirs(save_dir, exist_ok=True)

                for uploaded_file in uploaded_files:
                    ext = os.path.splitext(uploaded_file.name)[1]
                    saved_name = f"{uuid.uuid4().hex}{ext}"
                    saved_path = os.path.join(save_dir, saved_name)

                    with open(saved_path, 'wb+') as destination:
                        for chunk in uploaded_file.chunks():
                            destination.write(chunk)

                    doc_kwargs = {
                        'gift_request': obj,
                    }

                    if 'document_type' in doc_field_names:
                        doc_kwargs['document_type'] = 'STAFF_UPLOAD'

                    if 'original_file_name' in doc_field_names:
                        doc_kwargs['original_file_name'] = uploaded_file.name

                    if 'file_path' in doc_field_names:
                        doc_kwargs['file_path'] = saved_path

                    if 'mime_type' in doc_field_names:
                        guessed_type, _ = mimetypes.guess_type(uploaded_file.name)
                        doc_kwargs['mime_type'] = guessed_type or uploaded_file.content_type or 'application/octet-stream'

                    GiftRequestDocument.objects.create(**doc_kwargs)

                create_audit_log(
                    module_name='Gift Request',
                    table_name='gift_requests_giftrequest',
                    record_id=obj.id,
                    action_type='CREATE_REQUEST',
                    changed_by_type='STAFF',
                    changed_by_id=staff.id,
                    branch=obj.branch,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    remarks=remarks or 'Request created by staff',
                    field_changes=[
                        {'column_name': 'request_no', 'old_value': None, 'new_value': obj.request_no},
                        {'column_name': 'request_status', 'old_value': None, 'new_value': 'PENDING'},
                        {'column_name': 'updated_by_type', 'old_value': None, 'new_value': 'STAFF'},
                        {'column_name': 'updated_by_id', 'old_value': None, 'new_value': staff.id},
                    ]
                )

                messages.success(request, f'Request created successfully. Request No: {obj.request_no}')
                return redirect('staff-web-request-detail', gift_request_id=obj.id)

        except Exception as e:
            messages.error(request, f'Failed to create request: {str(e)}')
            return redirect('staff-web-create-request-form', share_id=share.id)

    return render(request, 'staff/create_request_form.html', {
        'share': share,
        'staff_user': staff,
    })
    
@staff_login_required
@permission_required('can_bulk_tracking')
def staff_bulk_tracking_upload(request):
    staff = request.staff_user

    preview_rows = []
    errors = []
    headers_found = []
    summary = None

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')

        validation_error = validate_excel_upload(uploaded_file)
        if validation_error:
            messages.error(request, validation_error)
            return render(request, 'staff/bulk_tracking_upload.html', {
                'staff_user': staff,
                'preview_rows': preview_rows,
                'errors': errors,
                'headers_found': headers_found,
                'summary': summary,
            })

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            expected_headers = ['request_no', 'courier_name', 'tracking_number', 'remarks']
            headers_found, header_error = validate_excel_headers_and_rows(ws, expected_headers)
            
            if header_error:
                errors.append(header_error)
            else:
                seen_request_nos = set()
                success_count = 0
                failed_count = 0
                failed_rows_for_export = []

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    request_no = (str(row[0]).strip() if row[0] is not None else '')
                    courier_name = (str(row[1]).strip() if row[1] is not None else '')
                    tracking_number = (str(row[2]).strip() if row[2] is not None else '')
                    remarks = (str(row[3]).strip() if row[3] is not None else '')

                    row_errors = []
                    obj = None

                    if not request_no:
                        row_errors.append('request_no is required')
                    if not courier_name:
                        row_errors.append('courier_name is required')
                    if not tracking_number:
                        row_errors.append('tracking_number is required')

                    if request_no:
                        if request_no in seen_request_nos:
                            row_errors.append('duplicate request_no in file')
                        else:
                            seen_request_nos.add(request_no)

                    if not row_errors and request_no:
                        obj = GiftRequest.objects.select_related('branch').filter(request_no=request_no).first()

                        if not obj:
                            row_errors.append('request not found')
                        else:
                            if not staff_has_branch_access(staff, obj.branch_id):
                                row_errors.append('no branch access')

                            if obj.request_status != 'ACCEPTED':
                                row_errors.append(f'only ACCEPTED requests allowed, current status is {obj.request_status}')

                    is_valid = len(row_errors) == 0

                    if is_valid and obj:
                        try:
                            with transaction.atomic():
                                old_status = obj.request_status
                                old_courier_name = obj.courier_name
                                old_tracking_number = obj.tracking_number
                                old_shipped_at = obj.shipped_at
                                old_updated_by_type = obj.updated_by_type
                                old_updated_by_id = obj.updated_by_id

                                obj.request_status = 'SHIPPED'
                                obj.courier_name = courier_name
                                obj.tracking_number = tracking_number
                                obj.shipped_at = timezone.now()
                                obj.updated_by_type = 'STAFF'
                                obj.updated_by_id = staff.id
                                obj.save()

                                GiftRequestStatusHistory.objects.create(
                                    gift_request=obj,
                                    old_status=old_status,
                                    new_status='SHIPPED',
                                    changed_by_type='STAFF',
                                    changed_by_id=staff.id,
                                    remarks=remarks or None
                                )

                                create_audit_log(
                                    module_name='Gift Request',
                                    table_name='gift_requests_giftrequest',
                                    record_id=obj.id,
                                    action_type='BULK_TRACKING_UPLOAD',
                                    changed_by_type='STAFF',
                                    changed_by_id=staff.id,
                                    branch=obj.branch,
                                    ip_address=request.META.get('REMOTE_ADDR'),
                                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                                    remarks=remarks or 'Bulk tracking upload by staff',
                                    field_changes=[
                                        {'column_name': 'request_status', 'old_value': old_status, 'new_value': obj.request_status},
                                        {'column_name': 'courier_name', 'old_value': old_courier_name, 'new_value': obj.courier_name},
                                        {'column_name': 'tracking_number', 'old_value': old_tracking_number, 'new_value': obj.tracking_number},
                                        {'column_name': 'shipped_at', 'old_value': old_shipped_at, 'new_value': obj.shipped_at},
                                        {'column_name': 'updated_by_type', 'old_value': old_updated_by_type, 'new_value': obj.updated_by_type},
                                        {'column_name': 'updated_by_id', 'old_value': old_updated_by_id, 'new_value': obj.updated_by_id},
                                    ]
                                )

                                send_tracking_created_sms(obj)

                            success_count += 1

                        except Exception as e:
                            row_errors.append(f'update failed: {str(e)}')
                            is_valid = False
                            failed_count += 1
                    else:
                        failed_count += 1

                    preview_rows.append({
                        'row_num': row_num,
                        'request_no': request_no,
                        'courier_name': courier_name,
                        'tracking_number': tracking_number,
                        'remarks': remarks,
                        'row_errors': row_errors,
                        'is_valid': is_valid,
                    })
                    if not is_valid:
                        failed_rows_for_export.append({
                            'row_num': row_num,
                            'request_no': request_no,
                            'courier_name': courier_name,
                            'tracking_number': tracking_number,
                            'remarks': remarks,
                            'row_errors': row_errors,
                        })

                summary = {
                    'total_rows': len(preview_rows),
                    'success_rows': success_count,
                    'failed_rows': failed_count,
                }
                request.session['bulk_tracking_failed_rows'] = failed_rows_for_export

                if success_count:
                    messages.success(request, f'{success_count} request(s) shipped successfully')

        except Exception as e:
            errors.append(f'Failed to read Excel file: {str(e)}')

    return render(request, 'staff/bulk_tracking_upload.html', {
        'staff_user': staff,
        'preview_rows': preview_rows,
        'errors': errors,
        'headers_found': headers_found,
        'summary': summary,
    })
    
@staff_login_required
@permission_required('can_bulk_tracking')
def staff_bulk_tracking_failed_export(request):
    failed_rows = request.session.get('bulk_tracking_failed_rows', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    headers = ['row_num', 'request_no', 'courier_name', 'tracking_number', 'remarks', 'error_message']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in failed_rows:
        ws.append([
            row.get('row_num', ''),
            row.get('request_no', ''),
            row.get('courier_name', ''),
            row.get('tracking_number', ''),
            row.get('remarks', ''),
            ', '.join(row.get('row_errors', [])) if row.get('row_errors') else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_tracking_failed_rows.xlsx'
    wb.save(response)
    return response

@staff_login_required
@permission_required('can_bulk_tracking')
def staff_bulk_tracking_sample(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Tracking Upload"

    headers = ['request_no', 'courier_name', 'tracking_number', 'remarks']
    ws.append(headers)

    sample_rows = [
        ['REQ0001', 'Blue Dart', 'BD123456789', 'Bulk upload test'],
        ['REQ0002', 'DTDC', 'DT987654321', 'Urgent shipment'],
    ]

    for row in sample_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_tracking_sample.xlsx'
    wb.save(response)
    return response
    
@staff_login_required
@permission_required('can_bulk_delivery')
def staff_bulk_delivery_upload(request):
    staff = request.staff_user

    preview_rows = []
    errors = []
    headers_found = []
    summary = None

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')

        validation_error = validate_excel_upload(uploaded_file)
        if validation_error:
            messages.error(request, validation_error)
            return render(request, 'staff/bulk_delivery_upload.html', {
                'staff_user': staff,
                'preview_rows': preview_rows,
                'errors': errors,
                'headers_found': headers_found,
                'summary': summary,
            })

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            expected_headers = ['request_no', 'delivery_date', 'remarks']
            headers_found, header_error = validate_excel_headers_and_rows(ws, expected_headers)

            if header_error:
                errors.append(header_error)
            else:
                seen_request_nos = set()
                success_count = 0
                failed_count = 0
                failed_rows_for_export = []

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    request_no = (str(row[0]).strip() if row[0] is not None else '')
                    delivery_date_raw = row[1]
                    remarks = (str(row[2]).strip() if row[2] is not None else '')

                    row_errors = []
                    obj = None
                    delivered_at_value = None

                    if not request_no:
                        row_errors.append('request_no is required')

                    if request_no:
                        if request_no in seen_request_nos:
                            row_errors.append('duplicate request_no in file')
                        else:
                            seen_request_nos.add(request_no)

                    if delivery_date_raw:
                        if hasattr(delivery_date_raw, 'year'):
                            try:
                                delivered_at_value = timezone.datetime(
                                    delivery_date_raw.year,
                                    delivery_date_raw.month,
                                    delivery_date_raw.day,
                                    0, 0, 0,
                                    tzinfo=timezone.get_current_timezone()
                                )
                            except Exception:
                                row_errors.append('invalid delivery_date')
                        else:
                            row_errors.append('delivery_date must be a valid Excel date')
                    else:
                        delivered_at_value = timezone.now()

                    if not row_errors and request_no:
                        obj = GiftRequest.objects.select_related('branch').filter(request_no=request_no).first()

                        if not obj:
                            row_errors.append('request not found')
                        else:
                            if not staff_has_branch_access(staff, obj.branch_id):
                                row_errors.append('no branch access')

                            if obj.request_status != 'SHIPPED':
                                row_errors.append(f'only SHIPPED requests allowed, current status is {obj.request_status}')

                    is_valid = len(row_errors) == 0

                    if is_valid and obj:
                        try:
                            with transaction.atomic():
                                old_status = obj.request_status
                                old_delivered_at = obj.delivered_at
                                old_updated_by_type = obj.updated_by_type
                                old_updated_by_id = obj.updated_by_id

                                obj.request_status = 'DELIVERED'
                                obj.delivered_at = delivered_at_value
                                obj.updated_by_type = 'STAFF'
                                obj.updated_by_id = staff.id
                                obj.save()

                                GiftRequestStatusHistory.objects.create(
                                    gift_request=obj,
                                    old_status=old_status,
                                    new_status='DELIVERED',
                                    changed_by_type='STAFF',
                                    changed_by_id=staff.id,
                                    remarks=remarks or None
                                )

                                create_audit_log(
                                    module_name='Gift Request',
                                    table_name='gift_requests_giftrequest',
                                    record_id=obj.id,
                                    action_type='BULK_DELIVERY_UPLOAD',
                                    changed_by_type='STAFF',
                                    changed_by_id=staff.id,
                                    branch=obj.branch,
                                    ip_address=request.META.get('REMOTE_ADDR'),
                                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                                    remarks=remarks or 'Bulk delivery upload by staff',
                                    field_changes=[
                                        {'column_name': 'request_status', 'old_value': old_status, 'new_value': obj.request_status},
                                        {'column_name': 'delivered_at', 'old_value': old_delivered_at, 'new_value': obj.delivered_at},
                                        {'column_name': 'updated_by_type', 'old_value': old_updated_by_type, 'new_value': obj.updated_by_type},
                                        {'column_name': 'updated_by_id', 'old_value': old_updated_by_id, 'new_value': obj.updated_by_id},
                                    ]
                                )

                            success_count += 1

                        except Exception as e:
                            row_errors.append(f'update failed: {str(e)}')
                            is_valid = False
                            failed_count += 1
                    else:
                        failed_count += 1

                    preview_rows.append({
                        'row_num': row_num,
                        'request_no': request_no,
                        'delivery_date': delivery_date_raw,
                        'remarks': remarks,
                        'row_errors': row_errors,
                        'is_valid': is_valid,
                    })

                    if not is_valid:
                        failed_rows_for_export.append({
                            'row_num': row_num,
                            'request_no': request_no,
                            'delivery_date': str(delivery_date_raw) if delivery_date_raw else '',
                            'remarks': remarks,
                            'row_errors': row_errors,
                        })

                summary = {
                    'total_rows': len(preview_rows),
                    'success_rows': success_count,
                    'failed_rows': failed_count,
                }

                request.session['bulk_delivery_failed_rows'] = failed_rows_for_export

                if success_count:
                    messages.success(request, f'{success_count} request(s) delivered successfully')

        except Exception as e:
            errors.append(f'Failed to read Excel file: {str(e)}')

    return render(request, 'staff/bulk_delivery_upload.html', {
        'staff_user': staff,
        'preview_rows': preview_rows,
        'errors': errors,
        'headers_found': headers_found,
        'summary': summary,
    })


@staff_login_required
@permission_required('can_bulk_delivery')
def staff_bulk_delivery_sample(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Delivery Upload"

    headers = ['request_no', 'delivery_date', 'remarks']
    ws.append(headers)

    sample_rows = [
        ['REQ0001', timezone.now().date(), 'Delivered successfully'],
        ['REQ0002', timezone.now().date(), 'Received by customer'],
    ]

    for row in sample_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_delivery_sample.xlsx'
    wb.save(response)
    return response


@staff_login_required
@permission_required('can_bulk_delivery')
def staff_bulk_delivery_failed_export(request):
    failed_rows = request.session.get('bulk_delivery_failed_rows', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    headers = ['row_num', 'request_no', 'delivery_date', 'remarks', 'error_message']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in failed_rows:
        ws.append([
            row.get('row_num', ''),
            row.get('request_no', ''),
            row.get('delivery_date', ''),
            row.get('remarks', ''),
            ', '.join(row.get('row_errors', [])) if row.get('row_errors') else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_delivery_failed_rows.xlsx'
    wb.save(response)
    return response
    
@staff_login_required
@permission_required('can_bulk_share_upload')
def staff_bulk_share_upload(request):
    staff = request.staff_user

    preview_rows = []
    errors = []
    headers_found = []
    summary = None

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')

        validation_error = validate_excel_upload(uploaded_file)
        if validation_error:
            messages.error(request, validation_error)
            return render(request, 'staff/bulk_share_upload.html', {
                'staff_user': staff,
                'preview_rows': preview_rows,
                'errors': errors,
                'headers_found': headers_found,
                'summary': summary,
            })

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            expected_headers = [
                'branch_id',
                'shareholder_code',
                'shareholder_name',
                'mobile_number',
                'email',
                'address_line1',
                'address_line2',
                'city',
                'state',
                'pincode',
                'share_number',
                'certificate_number',
                'gift_status',
                'stop_reason',
            ]
            headers_found, header_error = validate_excel_headers_and_rows(ws, expected_headers)

            if header_error:
                errors.append(header_error)
            else:
                seen_share_numbers = set()
                valid_count = 0
                invalid_count = 0
                valid_rows_for_import = []

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    branch_id = str(row[0]).strip() if row[0] is not None else ''
                    shareholder_code = str(row[1]).strip() if row[1] is not None else ''
                    shareholder_name = str(row[2]).strip() if row[2] is not None else ''
                    mobile_number = str(row[3]).strip() if row[3] is not None else ''
                    email = str(row[4]).strip() if row[4] is not None else ''
                    address_line1 = str(row[5]).strip() if row[5] is not None else ''
                    address_line2 = str(row[6]).strip() if row[6] is not None else ''
                    city = str(row[7]).strip() if row[7] is not None else ''
                    state = str(row[8]).strip() if row[8] is not None else ''
                    pincode = str(row[9]).strip() if row[9] is not None else ''
                    share_number = str(row[10]).strip() if row[10] is not None else ''
                    certificate_number = str(row[11]).strip() if row[11] is not None else ''
                    gift_status = str(row[12]).strip() if row[12] is not None else ''
                    stop_reason = str(row[13]).strip() if row[13] is not None else ''

                    row_errors = []
                    branch_obj = None

                    if not branch_id:
                        row_errors.append('branch_id is required')
                    if not shareholder_code:
                        row_errors.append('shareholder_code is required')
                    if not shareholder_name:
                        row_errors.append('shareholder_name is required')
                    if not mobile_number:
                        row_errors.append('mobile_number is required')
                    if not share_number:
                        row_errors.append('share_number is required')
                    if not gift_status:
                        row_errors.append('gift_status is required')

                    if gift_status and gift_status not in ['ELIGIBLE', 'STOPPED']:
                        row_errors.append('gift_status must be ELIGIBLE or STOPPED')

                    if gift_status == 'STOPPED' and not stop_reason:
                        row_errors.append('stop_reason is required when gift_status is STOPPED')

                    if branch_id:
                        if not str(branch_id).isdigit():
                            row_errors.append('branch_id must be numeric')
                        else:
                            branch_obj = Branch.objects.filter(id=int(branch_id), is_active=True).first()
                            if not branch_obj:
                                row_errors.append('branch not found')
                            elif not staff_has_branch_access(staff, branch_obj.id):
                                row_errors.append('no branch access')

                    if share_number:
                        if share_number in seen_share_numbers:
                            row_errors.append('duplicate share_number in file')
                        else:
                            seen_share_numbers.add(share_number)

                        if Share.objects.filter(share_number=share_number).exists():
                            row_errors.append('share_number already exists in database')

                    if certificate_number and Share.objects.filter(certificate_number=certificate_number).exists():
                        row_errors.append('certificate_number already exists in database')

                    is_valid = len(row_errors) == 0

                    row_payload = {
                        'row_num': row_num,
                        'branch_id': branch_id,
                        'shareholder_code': shareholder_code,
                        'shareholder_name': shareholder_name,
                        'mobile_number': mobile_number,
                        'email': email,
                        'address_line1': address_line1,
                        'address_line2': address_line2,
                        'city': city,
                        'state': state,
                        'pincode': pincode,
                        'share_number': share_number,
                        'certificate_number': certificate_number,
                        'gift_status': gift_status,
                        'stop_reason': stop_reason,
                        'row_errors': row_errors,
                        'is_valid': is_valid,
                    }

                    preview_rows.append(row_payload)

                    if is_valid:
                        valid_count += 1
                        valid_rows_for_import.append(row_payload)
                    else:
                        invalid_count += 1

                summary = {
                    'total_rows': len(preview_rows),
                    'valid_rows': valid_count,
                    'invalid_rows': invalid_count,
                }

                request.session['bulk_share_valid_rows'] = valid_rows_for_import

        except Exception as e:
            errors.append(f'Failed to read Excel file: {str(e)}')

    return render(request, 'staff/bulk_share_upload.html', {
        'staff_user': staff,
        'preview_rows': preview_rows,
        'errors': errors,
        'headers_found': headers_found,
        'summary': summary,
    })
    
@staff_login_required
@permission_required('can_bulk_share_upload')
def staff_bulk_share_confirm_import(request):
    staff = request.staff_user

    if request.method != 'POST':
        return redirect('staff-web-bulk-share-upload')

    valid_rows = request.session.get('bulk_share_valid_rows', [])

    if not valid_rows:
        messages.error(request, 'No validated rows found for import')
        return redirect('staff-web-bulk-share-upload')

    success_count = 0
    failed_count = 0
    failed_rows = []

    for row in valid_rows:
        try:
            with transaction.atomic():
                branch_obj = Branch.objects.get(id=int(row['branch_id']), is_active=True)

                if not staff_has_branch_access(staff, branch_obj.id):
                    raise Exception('No branch access')

                if Share.objects.filter(share_number=row['share_number']).exists():
                    raise Exception('share_number already exists in database')

                if row['certificate_number'] and Share.objects.filter(certificate_number=row['certificate_number']).exists():
                    raise Exception('certificate_number already exists in database')

                shareholder = Shareholder.objects.filter(
                    shareholder_code=row['shareholder_code']
                ).first()

                if shareholder:
                    shareholder.shareholder_name = row['shareholder_name']
                    shareholder.mobile_number = row['mobile_number']
                    shareholder.email = row['email'] or None
                    shareholder.address_line1 = row['address_line1'] or None
                    shareholder.address_line2 = row['address_line2'] or None
                    shareholder.city = row['city'] or None
                    shareholder.state = row['state'] or None
                    shareholder.pincode = row['pincode'] or None
                    shareholder.is_active = True
                    shareholder.save()
                else:
                    shareholder = Shareholder.objects.create(
                        shareholder_code=row['shareholder_code'],
                        shareholder_name=row['shareholder_name'],
                        mobile_number=row['mobile_number'],
                        email=row['email'] or None,
                        address_line1=row['address_line1'] or None,
                        address_line2=row['address_line2'] or None,
                        city=row['city'] or None,
                        state=row['state'] or None,
                        pincode=row['pincode'] or None,
                        is_active=True,
                    )

                share = Share.objects.create(
                    shareholder=shareholder,
                    share_number=row['share_number'],
                    certificate_number=row['certificate_number'] or None,
                    branch=branch_obj,
                    gift_status=row['gift_status'],
                    stop_reason=row['stop_reason'] or None,
                    is_active=True,
                )

                create_audit_log(
                    module_name='Share Master',
                    table_name='shares_share',
                    record_id=share.id,
                    action_type='BULK_SHARE_IMPORT',
                    changed_by_type='STAFF',
                    changed_by_id=staff.id,
                    branch=branch_obj,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    remarks='Bulk shareholder/share import by staff',
                    field_changes=[
                        {'column_name': 'shareholder_code', 'old_value': None, 'new_value': shareholder.shareholder_code},
                        {'column_name': 'shareholder_name', 'old_value': None, 'new_value': shareholder.shareholder_name},
                        {'column_name': 'share_number', 'old_value': None, 'new_value': share.share_number},
                        {'column_name': 'certificate_number', 'old_value': None, 'new_value': share.certificate_number},
                        {'column_name': 'gift_status', 'old_value': None, 'new_value': share.gift_status},
                    ]
                )

                success_count += 1

        except Exception as e:
            failed_count += 1
            failed_rows.append({
                'row_num': row.get('row_num', ''),
                'branch_id': row.get('branch_id', ''),
                'shareholder_code': row.get('shareholder_code', ''),
                'shareholder_name': row.get('shareholder_name', ''),
                'mobile_number': row.get('mobile_number', ''),
                'share_number': row.get('share_number', ''),
                'certificate_number': row.get('certificate_number', ''),
                'gift_status': row.get('gift_status', ''),
                'stop_reason': row.get('stop_reason', ''),
                'row_errors': [str(e)],
            })

    request.session['bulk_share_failed_rows'] = failed_rows
    request.session.pop('bulk_share_valid_rows', None)

    if success_count:
        messages.success(request, f'{success_count} shareholder/share row(s) imported successfully')

    if failed_count:
        messages.error(request, f'{failed_count} row(s) failed during import')

    return redirect('staff-web-bulk-share-upload-result')
    
@staff_login_required
@permission_required('can_bulk_share_upload')
def staff_bulk_share_upload_result(request):
    staff = request.staff_user
    failed_rows = request.session.get('bulk_share_failed_rows', [])

    summary = {
        'total_rows': len(failed_rows),
        'valid_rows': 0,
        'invalid_rows': len(failed_rows),
    }

    return render(request, 'staff/bulk_share_upload_result.html', {
        'staff_user': staff,
        'failed_rows': failed_rows,
        'summary': summary,
    })
    
@staff_login_required
@permission_required('can_bulk_share_upload')
def staff_bulk_share_failed_export(request):
    failed_rows = request.session.get('bulk_share_failed_rows', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    headers = [
        'row_num',
        'branch_id',
        'shareholder_code',
        'shareholder_name',
        'mobile_number',
        'share_number',
        'certificate_number',
        'gift_status',
        'stop_reason',
        'error_message',
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in failed_rows:
        ws.append([
            row.get('row_num', ''),
            row.get('branch_id', ''),
            row.get('shareholder_code', ''),
            row.get('shareholder_name', ''),
            row.get('mobile_number', ''),
            row.get('share_number', ''),
            row.get('certificate_number', ''),
            row.get('gift_status', ''),
            row.get('stop_reason', ''),
            ', '.join(row.get('row_errors', [])) if row.get('row_errors') else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_share_failed_rows.xlsx'
    wb.save(response)
    return response

@staff_login_required
@permission_required('can_bulk_share_upload')
def staff_bulk_share_sample(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Share Upload"

    headers = [
        'branch_id',
        'shareholder_code',
        'shareholder_name',
        'mobile_number',
        'email',
        'address_line1',
        'address_line2',
        'city',
        'state',
        'pincode',
        'share_number',
        'certificate_number',
        'gift_status',
        'stop_reason',
    ]
    ws.append(headers)

    sample_rows = [
        [
            1,
            'SH0001',
            'Ramesh Patel',
            '9876543210',
            'ramesh@example.com',
            'Street 1',
            'Near Market',
            'Rajkot',
            'Gujarat',
            '360001',
            'SHARE1001',
            'CERT1001',
            'ELIGIBLE',
            '',
        ],
        [
            1,
            'SH0002',
            'Suresh Mehta',
            '9876500000',
            'suresh@example.com',
            'Street 2',
            '',
            'Rajkot',
            'Gujarat',
            '360002',
            'SHARE1002',
            'CERT1002',
            'STOPPED',
            'Document pending',
        ],
    ]

    for row in sample_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_share_upload_sample.xlsx'
    wb.save(response)
    return response
    
@staff_login_required
@permission_required('can_bulk_share_status')
def staff_bulk_share_status_upload(request):
    staff = request.staff_user

    preview_rows = []
    errors = []
    headers_found = []
    summary = None

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')

        validation_error = validate_excel_upload(uploaded_file)
        if validation_error:
            messages.error(request, validation_error)
            return render(request, 'staff/bulk_share_status_upload.html', {
                'staff_user': staff,
                'preview_rows': preview_rows,
                'errors': errors,
                'headers_found': headers_found,
                'summary': summary,
            })

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            expected_headers = ['share_number', 'gift_status', 'remarks']
            headers_found, header_error = validate_excel_headers_and_rows(ws, expected_headers)

            if header_error:
                errors.append(header_error)
            else:
                seen_share_numbers = set()
                success_count = 0
                failed_count = 0
                failed_rows_for_export = []

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    share_number = str(row[0]).strip() if row[0] is not None else ''
                    gift_status = str(row[1]).strip() if row[1] is not None else ''
                    remarks = str(row[2]).strip() if row[2] is not None else ''

                    row_errors = []
                    share_obj = None

                    if not share_number:
                        row_errors.append('share_number is required')

                    if not gift_status:
                        row_errors.append('gift_status is required')

                    if gift_status and gift_status not in ['ELIGIBLE', 'STOPPED']:
                        row_errors.append('gift_status must be ELIGIBLE or STOPPED')

                    if not remarks:
                        row_errors.append('remarks is required')

                    if share_number:
                        if share_number in seen_share_numbers:
                            row_errors.append('duplicate share_number in file')
                        else:
                            seen_share_numbers.add(share_number)

                    if not row_errors and share_number:
                        share_obj = Share.objects.select_related('branch').filter(
                            share_number=share_number,
                            is_active=True
                        ).first()

                        if not share_obj:
                            row_errors.append('share not found')
                        else:
                            if not staff_has_branch_access(staff, share_obj.branch_id):
                                row_errors.append('no branch access')

                    is_valid = len(row_errors) == 0

                    if is_valid and share_obj:
                        try:
                            with transaction.atomic():
                                old_status = share_obj.gift_status
                                old_reason = share_obj.stop_reason

                                share_obj.gift_status = gift_status
                                share_obj.stop_reason = remarks
                                share_obj.save()

                                create_audit_log(
                                    module_name='Share Master',
                                    table_name='shares_share',
                                    record_id=share_obj.id,
                                    action_type='BULK_SHARE_STATUS_UPDATE',
                                    changed_by_type='STAFF',
                                    changed_by_id=staff.id,
                                    branch=share_obj.branch,
                                    ip_address=request.META.get('REMOTE_ADDR'),
                                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                                    remarks=remarks,
                                    field_changes=[
                                        {'column_name': 'gift_status', 'old_value': old_status, 'new_value': share_obj.gift_status},
                                        {'column_name': 'stop_reason', 'old_value': old_reason, 'new_value': share_obj.stop_reason},
                                    ]
                                )

                            success_count += 1

                        except Exception as e:
                            row_errors.append(f'update failed: {str(e)}')
                            is_valid = False
                            failed_count += 1
                    else:
                        failed_count += 1

                    preview_rows.append({
                        'row_num': row_num,
                        'share_number': share_number,
                        'gift_status': gift_status,
                        'remarks': remarks,
                        'row_errors': row_errors,
                        'is_valid': is_valid,
                    })

                    if not is_valid:
                        failed_rows_for_export.append({
                            'row_num': row_num,
                            'share_number': share_number,
                            'gift_status': gift_status,
                            'remarks': remarks,
                            'row_errors': row_errors,
                        })

                summary = {
                    'total_rows': len(preview_rows),
                    'success_rows': success_count,
                    'failed_rows': failed_count,
                }

                request.session['bulk_share_status_failed_rows'] = failed_rows_for_export

                if success_count:
                    messages.success(request, f'{success_count} share status row(s) updated successfully')

        except Exception as e:
            errors.append(f'Failed to read Excel file: {str(e)}')

    return render(request, 'staff/bulk_share_status_upload.html', {
        'staff_user': staff,
        'preview_rows': preview_rows,
        'errors': errors,
        'headers_found': headers_found,
        'summary': summary,
    })


@staff_login_required
@permission_required('can_bulk_share_status')
def staff_bulk_share_status_sample(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Share Status"

    headers = ['share_number', 'gift_status', 'remarks']
    ws.append(headers)

    sample_rows = [
        ['SHARE1001', 'STOPPED', 'Documents pending'],
        ['SHARE1002', 'ELIGIBLE', 'Issue resolved'],
    ]

    for row in sample_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_share_status_sample.xlsx'
    wb.save(response)
    return response


@staff_login_required
@permission_required('can_bulk_share_status')
def staff_bulk_share_status_failed_export(request):
    failed_rows = request.session.get('bulk_share_status_failed_rows', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Rows"

    headers = ['row_num', 'share_number', 'gift_status', 'remarks', 'error_message']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in failed_rows:
        ws.append([
            row.get('row_num', ''),
            row.get('share_number', ''),
            row.get('gift_status', ''),
            row.get('remarks', ''),
            ', '.join(row.get('row_errors', [])) if row.get('row_errors') else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_share_status_failed_rows.xlsx'
    wb.save(response)
    return response